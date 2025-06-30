#!/usr/bin/env python3
import re
import pandas as pd
import numpy as np
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import LineChart, Reference

# ── CONFIGURE THESE TO YOUR LOCAL PATHS ───────────────────────────────────────
INPUT_CSV   = r"C:\Users\jat123\Documents\Analytics R\SMEC 666 Echo Data.csv"
OUTPUT_XLSX = r"C:\Users\jat123\Documents\Analytics R\SMEC 666 Full.xlsx"
# ──────────────────────────────────────────────────────────────────────────────

def time_to_seconds(ts: str) -> int:
    if pd.isna(ts) or ts == "":
        return 0
    parts = list(map(int, ts.split(":")))
    while len(parts) < 3:
        parts.insert(0, 0)
    h, m, s = parts
    return h*3600 + m*60 + s

def seconds_to_hms(sec: float) -> str:
    return "" if pd.isna(sec) else str(timedelta(seconds=int(sec)))

def natural_key(s: str):
    return [int(chunk) if chunk.isdigit() else chunk.lower()
            for chunk in re.split(r'(\d+)', s)]

def main():
    # 1) Load & parse
    df = pd.read_csv(INPUT_CSV, dtype=str)
    df['Duration_sec']      = df['Duration'].apply(time_to_seconds)
    df['TotalViewTime_sec'] = df['Total View Time'].apply(time_to_seconds)
    df['AvgViewTime_sec']   = df['Average View Time'].apply(time_to_seconds)
    df['Row View %']        = df['TotalViewTime_sec'] / df['Duration_sec'].replace(0, np.nan)

    # 2) Group
    grp = df.groupby('Media Name', sort=False)
    titles = list(grp.groups.keys())
    media_count = len(titles)

    # 3) Build summary_core
    summary_core = pd.DataFrame({
        'Media Title':              titles,
        'Video Duration':           [grp.get_group(t)['Duration_sec'].iloc[0] for t in titles],
        'Number of Unique Viewers': grp['User Name'].nunique().values,
        'Average View %':           grp['Row View %'].mean().fillna(0).values,
        'Total View %':             (grp['TotalViewTime_sec'].sum()
                                     / grp['Duration_sec'].sum()).values,
        'Total View Time':          grp['TotalViewTime_sec'].sum().values,
        'Average View Time':        grp['AvgViewTime_sec'].mean().values,
        'Average Total View Time':  grp['TotalViewTime_sec'].mean().values,
    })

    # 4) Natural sort
    summary_core['sort_key'] = summary_core['Media Title'].apply(natural_key)
    summary_core = (
        summary_core
        .sort_values('sort_key')
        .drop(columns='sort_key')
        .reset_index(drop=True)
    )

    # 5) Grand Total (now averages)
    means = summary_core[[
        'Video Duration',
        'Total View Time',
        'Average View Time',
        'Average Total View Time'
    ]].mean()
    viewers_mean = summary_core['Number of Unique Viewers'].mean()
    summary_core.loc[len(summary_core)] = {
        'Media Title':               'Grand Total',
        'Video Duration':            means['Video Duration'],
        'Number of Unique Viewers':  viewers_mean,
        'Average View %':            summary_core['Average View %'].mean(),
        'Total View %':              summary_core['Total View %'].mean(),
        'Total View Time':           means['Total View Time'],
        'Average View Time':         means['Average View Time'],
        'Average Total View Time':   means['Average Total View Time'],
    }

    # 6) Average Video Length and Watch Time
    n = len(summary_core) - 1
    means2 = summary_core.loc[:n-1, [
        'Video Duration',
        'Total View Time',
        'Average View Time',
        'Average Total View Time'
    ]].mean()
    summary_core.loc[len(summary_core)] = {
        'Media Title':               'Average Video Length and Watch Time',
        'Video Duration':            means2['Video Duration'],
        'Number of Unique Viewers':  '',
        'Average View %':            summary_core.loc[:n-1, 'Average View %'].mean(),
        'Total View %':              summary_core.loc[:n-1, 'Total View %'].mean(),
        'Total View Time':           means2['Total View Time'],
        'Average View Time':         means2['Average View Time'],
        'Average Total View Time':   means2['Average Total View Time'],
    }

    # 7) Convert time cols to strings for initial write
    for col in [
        'Video Duration',
        'Total View Time',
        'Average View Time',
        'Average Total View Time'
    ]:
        summary_core[col] = summary_core[col].apply(seconds_to_hms)

    # 8) Reorder columns
    summary = summary_core[[
        'Media Title',
        'Video Duration',
        'Number of Unique Viewers',
        'Average View %',
        'Total View %',
        'Total View Time',
        'Average View Time',
        'Average Total View Time'
    ]]

    # 9) Write to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Media Summary'
    for row in dataframe_to_rows(summary, index=False, header=True):
        ws.append(row)
    last_row = ws.max_row

    # 10) Convert Video Duration (col B) back to numeric Excel time
    for r in range(2, last_row + 1):
        cell = ws[f'B{r}']
        secs = time_to_seconds(cell.value)
        cell.value = secs / 86400.0
        cell.number_format = 'hh:mm:ss'

    # 11) Format other columns
    for r in range(2, last_row + 1):
        # Percent columns: D, E
        for col in ('D','E'):
            c = ws[f'{col}{r}']
            if isinstance(c.value, (int, float)):
                c.number_format = '0.00%'
        # Time columns: F, G, H
        for col in ('F','G','H'):
            ws[f'{col}{r}'].number_format = 'hh:mm:ss'

    # 12) Add orange data bars on Video Duration (B) & Average View % (D)
    bar = DataBarRule(start_type='min', end_type='max', color="FFA500")
    ws.conditional_formatting.add(f"B2:B{media_count+1}", bar)
    ws.conditional_formatting.add(f"D2:D{media_count+1}", bar)

    # 13) Create line charts
    # Chart 1: Average View % Over Time
    chart1 = LineChart()
    chart1.title = "View % Over Time"
    chart1.style = 9
    chart1.y_axis.number_format = '0.00%'
    data1 = Reference(ws, min_col=4, min_row=1, max_row=media_count+1)
    chart1.add_data(data1, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=2, max_row=media_count+1)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "J2")

    # Chart 2: Unique Viewers Over Time
    chart2 = LineChart()
    chart2.title = "Unique Viewers Over Time"
    chart2.style = 9
    data2 = Reference(ws, min_col=3, min_row=1, max_row=media_count+1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "J20")

    # 14) Table styling
    tbl = Table(displayName="MediaStats", ref=f"A1:H{last_row}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    # 15) Save
    wb.save(OUTPUT_XLSX)
    print(f"✅ Written: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
