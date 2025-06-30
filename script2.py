import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# ----- CONFIGURE THESE PATHS -----
input_csv   = r"C:\Users\jat123\OneDrive - Colostate\Documents\Analytics\Analytics Pilot\SM25\HDFS 439 HalseyD\Grades.csv"
output_dir  = r"C:\Users\jat123\OneDrive - Colostate\Documents\Analytics\Analytics Pilot\SM25\HDFS 439 HalseyD"
output_file = os.path.join(output_dir, "processed_output.xlsx")
# ---------------------------------

# 1. Read CSV into DataFrame
df = pd.read_csv(input_csv)

# 2. Drop any row where the first column contains "Student, Test"
mask = df.iloc[:, 0].astype(str).str.contains("Student, Test", na=False)
df = df[~mask].reset_index(drop=True)

# 3. Drop unwanted columns if they exist (but NOT "Final Grade"):
to_drop = [
    "Student", "ID", "SIS User ID", "SIS Login ID",
    "Current Grade", "Unposted Current Grade", "Unposted Final Grade"
]
df.drop(columns=[c for c in to_drop if c in df.columns], inplace=True)

# 4. Drop any column where rows 3+ are all empty or only zeros, except "Final Grade"
drop_cols = []
for col in df.columns:
    if col == "Final Grade":
        continue
    s = pd.to_numeric(df[col].iloc[2:], errors='coerce')
    if s.fillna(0).eq(0).all():
        drop_cols.append(col)
df.drop(columns=drop_cols, inplace=True)

# 5. Ensure output folder exists
os.makedirs(output_dir, exist_ok=True)

# 6. Write intermediate data to Excel
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, index=False)

# 7. Re-open with openpyxl to prepare for in-place edits
wb = load_workbook(output_file)
ws = wb.active

# 7a. Find the column index of "Final Grade" (before inserting titles col)
final_grade_idx_pre = None
for ci in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=ci).value == "Final Grade":
        final_grade_idx_pre = ci
        break

# 7b. Fill all empty cells in columns 2+ with numeric 0, skipping Final Grade
for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                        min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.column == final_grade_idx_pre:
            continue
        if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
            cell.value = 0

# 7c. Convert all remaining strings→numbers where possible, skipping Final Grade
for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                        min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.column == final_grade_idx_pre:
            continue
        if isinstance(cell.value, str) and cell.value.strip():
            txt = cell.value.replace(",", "")
            try:
                cell.value = float(txt)
            except ValueError:
                pass

# 7d. Replace any "(read only)" in row 2 with that column’s max (rows 3+), skipping Final Grade
data_last_row = ws.max_row
for ci in range(1, ws.max_column + 1):
    if ci == final_grade_idx_pre:
        continue
    hdr = ws.cell(row=2, column=ci)
    if isinstance(hdr.value, str) and "(read only)" in hdr.value:
        nums = [
            ws.cell(row=r, column=ci).value
            for r in range(3, data_last_row + 1)
            if isinstance(ws.cell(row=r, column=ci).value, (int, float))
        ]
        if nums:
            hdr.value = max(nums)

# 8. Insert a new blank column A for Row Titles
ws.insert_cols(1)
ws["A1"] = "Row Titles"

#    Re-find "Final Grade" now that columns have shifted
final_grade_idx = None
for ci in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=ci).value == "Final Grade":
        final_grade_idx = ci
        break

# 9. Label row 2 as "Points Possible"
ws["A2"] = "Points Possible"

# 10. Append "Average" and "Average Excluding Zeros" rows
original_last_data_row = ws.max_row
avg_row  = original_last_data_row + 1
avg0_row = original_last_data_row + 2
ws[f"A{avg_row}"]  = "Average"
ws[f"A{avg0_row}"] = "Average Excluding Zeros"

# 11. Fill in Average formulas & % formatting (skip Final Grade)
max_col = ws.max_column
for col in range(2, max_col + 1):
    if col == final_grade_idx:
        continue
    letter   = get_column_letter(col)
    data_rng = f"{letter}3:{letter}{original_last_data_row}"
    header   = f"{letter}$2"
    c_avg    = ws[f"{letter}{avg_row}"]
    c_avg.value = f"=AVERAGE({data_rng})/{header}"
    c_avg.number_format = '0.00%'
    c_avg0   = ws[f"{letter}{avg0_row}"]
    c_avg0.value = f"=AVERAGEIF({data_rng},\">0\")/{header}"
    c_avg0.number_format = '0.00%'

# 12. Apply conditional formatting (>90%, 80–90%, <80%), skip Final Grade
green  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

for row in (avg_row, avg0_row):
    rng = f"B{row}:{get_column_letter(max_col)}{row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator='greaterThan', formula=['0.9'], fill=green)
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator='between',     formula=['0.8','0.9'], fill=yellow)
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator='lessThan',    formula=['0.8'], fill=red)
    )

# 13. Count & Percent of F in Final Grade
count_row = avg0_row + 1
pct_row   = avg0_row + 2
ws[f"A{count_row}"] = "Count of F"
ws[f"A{pct_row}"]   = "Percent of F"
fg_letter = get_column_letter(final_grade_idx)
ws.cell(row=count_row, column=final_grade_idx).value = (
    f'=COUNTIF({fg_letter}3:{fg_letter}{original_last_data_row},"F")'
)
total_students = original_last_data_row - 2
ws.cell(row=pct_row, column=final_grade_idx).value = (
    f'={fg_letter}{count_row}/{total_students}'
)
ws.cell(row=pct_row, column=final_grade_idx).number_format = '0.00%'

# 14. Format the data region as an Excel Table ("Dark Teal, Table Style Medium 2")
#     from A1 through the last column & the last student row
table_end = get_column_letter(max_col) + str(original_last_data_row)
table = Table(displayName="GradesTable", ref=f"A1:{table_end}")
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

# 15. Save the workbook
wb.save(output_file)
print(f"✅ Done – saved to {output_file}")